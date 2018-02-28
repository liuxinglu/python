#!/usr/bin/python
# -*- coding: UTF-8 -*-
# import os, sys

# file = "/Users/star_xlliu/save.txt"
# exist = os.access(file, os.F_OK) #是否存在
# readabled = os.access(file, os.R_OK) #是否可读
# writeabled = os.access(file, os.W_OK) #是否可写
# exeabled = os.access(file, os.X_OK) #是否可执行
# print exist, readabled, writeabled, exeabled

# path = "/Users/star_xlliu/Documents/work/tmp"
# retPath = os.getcwd() #当前工作目录
# print retPath
# try:
#     os.removedirs(path)
#     os.chdir(path)
# except OSError:
#     os.mkdir(path)
#     os.chdir(path)
#     print os.listdir(os.getcwd())


# try:
#     fh = open(file, "w")
#     fh.write("测试测试测试")
# except IOError:
#     print "写入错误"
# else:
#     print "写入成功"
#     fh.close()


import os
import xlrd
import openpyxl
import json
import codecs
from collections import OrderedDict

def xlsx2json_2(args):
    wb = xlrd.open_workbook(args)
    shArr = wb.sheet_names() #获取所有sheet名
    for shname in shArr:
        convert_list = []
        sh = wb.sheet_by_name(shname) #通过sheet名获得对应sheet内容
        title = sh.row_values(1)  #获取第二行的内容作为title
        for rownum in range(2, sh.nrows): #从第三行开始就是数据了
            rowvalue = sh.row_values(rownum) #每次拿一行数据
            single = OrderedDict()
            for colnum in range(1, len(rowvalue)): #将这行数据按列读取放入字典中
                single[title[colnum]] = rowvalue[colnum]
            convert_list.append(single) #读完之后将字段放入列表中

        j = json.dumps(convert_list, ensure_ascii=False)

        with codecs.open('/Users/star_xlliu/Documents/json/' + shname + '.json', "w", "utf-8") as f:
            f.write(j)
            f.close()

# xlsx2json_2('/Users/star_xlliu/Documents/test2.xlsx')

def readFromJson(path):
    with open(path, 'r') as f:
        jsonData = json.load(f)
    return jsonData

def jsonPath2xlsx(path):
    filePaths = os.listdir(path)
    filePaths = filePaths[1:]
    sheetArr = []
    excel = openpyxl.Workbook()
    for item in filePaths:
        sheet = excel.create_sheet(unicode(item, "utf-8")[: 6], index=0)
        sheetArr.append(sheet)

    index = 0
    for item in filePaths:
        
        data = readFromJson(path + item)
        length = len(data)
        i, j = 0, 0
        line0 = data[0]
        for k, v in line0.items():
            j += 1
            sheetArr[index].cell(row=1, column=j, value=k)
        while i < length:
            eachLine = data[i]
            j = 0
            i += 1;
            for v in eachLine.values():
                j += 1
                sheetArr[index].cell(row=i + 1, column=j, value=v)
        index += 1
    excel.save("/Users/star_xlliu/Documents/test3.xlsx")


jsonPath2xlsx('/Users/star_xlliu/Documents/json/');
